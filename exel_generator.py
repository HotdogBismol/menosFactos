import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.fills import PatternFill


class ExcelGenerator:
    def __init__(self):
        pass

    def generate(self, data_list, output_path):
        """Crea un Excel con formato condicional para alertas EFOS."""
        if not data_list:
            return False

        # --- Verificar que el archivo no esté abierto ---
        try:
            import os
            if os.path.exists(output_path):
                with open(output_path, 'a'):
                    pass  # solo comprueba acceso de escritura
        except PermissionError:
            raise PermissionError(
                f"El archivo está abierto en otro programa. Ciérralo e intenta de nuevo:\n{output_path}"
            )

        df = pd.DataFrame(data_list)

        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Reporte CFDI')

        worksheet = writer.sheets['Reporte CFDI']

        # --- Estilo de encabezados ---
        header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # --- Formato condicional: rojo para alertas EFOS ---
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        red_font = Font(color="9C0006")

        if "Alerta EFOS (Lista 69-B)" in df.columns:
            efos_col_idx = df.columns.get_loc("Alerta EFOS (Lista 69-B)") + 1

            for row_idx, row in enumerate(
                worksheet.iter_rows(
                    min_row=2, max_row=worksheet.max_row,
                    min_col=1, max_col=worksheet.max_column
                ),
                start=2
            ):
                alerta = worksheet.cell(row=row_idx, column=efos_col_idx).value
                if alerta == "SÍ":
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = red_font

        # --- Autoajuste de columnas (bug corregido) ---
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    # CORRECCIÓN: siempre convertir a str antes de len()
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
            # Limitar ancho máximo a 60 para evitar columnas enormes
            worksheet.column_dimensions[col_letter].width = min(max_length + 4, 60)

        writer.close()
        return True
